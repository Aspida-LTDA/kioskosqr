from PIL import Image, ImageDraw, ImageFont, ImageColor
import openpyxl 

file = "plantilla.png"
font = ImageFont.truetype("arial.ttf", 18)
img = Image.open(file)


def create_image(size, bgColor, message, font, fontColor):
    W, H = size
    image = Image.new('RGB', size, bgColor)
    draw = ImageDraw.Draw(image)
    _, _, w, h = draw.textbbox((0, 0), message, font=font)
    draw.text(((W-w)/2, (H-h)/2), message, font=font, fill=fontColor)
    return image

draw = ImageDraw.Draw(img)

spacing = 1
#myImage = create_image((200, 50), 'white', cedula, font, 'black')
#myImage.show()
#draw.text((45, 192), cedula, fill ="black", font = font,  spacing = spacing, align ="center") 
#back_im = img.copy()
#back_im.paste(im2, (75, 32))
#dataframe1.max_row
dataframe = openpyxl.load_workbook("ice.xlsx")
dataframe1 = dataframe.active
for row in range(3, dataframe1.max_row):
    imagennumber = row-2
    imagentexto = ""
    if imagennumber<10:
        imagentexto = "00" + str(imagennumber)
    elif imagennumber<100:
        imagentexto = "0" + str(imagennumber)
    else:
        imagentexto = str(imagennumber)
    cedula = str(dataframe1.cell(row, 1).value) + """
    """ + dataframe1.cell(row, 5).value + """
    """ + dataframe1.cell(row, 6).value
    archivo = str(dataframe1.cell(row, 1).value) + ".png"
    print(imagentexto)
    im2 = Image.open('imagenes\image'+str(imagentexto)+'.png')
    back_im = img.copy()
    back_im.paste(im2, (75, 32))
    draw = ImageDraw.Draw(back_im)
    draw.text((45, 192), cedula, fill ="black", font = font,  spacing = spacing, align ="center") 
    back_im.save(archivo)
    


#back_im.show()
#img.save("test.png")